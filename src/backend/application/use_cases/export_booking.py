"""Use case for exporting a single booking to Excel."""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.ports.output.repositories import BookingRepository, CompanyRepository


class ExportBookingUseCase:
    """Export booking detail and charges as an XLSX file."""

    def __init__(
        self,
        booking_repo: BookingRepository,
        company_repo: CompanyRepository,
    ) -> None:
        self.booking_repo = booking_repo
        self.company_repo = company_repo

    def execute(self, bl_reference: str) -> bytes:
        """Generate XLSX bytes for the requested booking."""
        booking = self.booking_repo.find_by_id(bl_reference)
        if booking is None:
            raise ValueError(f"Booking '{bl_reference}' not found")

        company = self.company_repo.get()
        commission_rate = company.agent_commission_rate if company else Decimal("0.50")

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Booking Export"

        header_rows = [
            ("Booking ID", booking.id),
            ("Status", booking.status.value),
            ("Client", booking.client.name if booking.client else ""),
            ("Client NIF", booking.client.nif if booking.client else ""),
            ("POL", booking.pol.code if booking.pol else ""),
            ("POD", booking.pod.code if booking.pod else ""),
            ("Vessel", booking.vessel or ""),
            ("Total Revenue", float(booking.total_revenue.amount)),
            ("Total Costs", float(booking.total_costs.amount)),
            ("Margin", float(booking.margin.amount)),
            (
                "Commission",
                float(booking.calculate_agent_commission(commission_rate).amount),
            ),
        ]
        for label, value in header_rows:
            sheet.append([label, value])
        for row in sheet.iter_rows(min_row=1, max_row=len(header_rows), min_col=1, max_col=1):
            row[0].font = Font(bold=True)

        start_row = len(header_rows) + 3
        sheet.cell(row=start_row, column=1, value="Type")
        sheet.cell(row=start_row, column=2, value="Category")
        sheet.cell(row=start_row, column=3, value="Description")
        sheet.cell(row=start_row, column=4, value="Container")
        sheet.cell(row=start_row, column=5, value="Amount")
        for column in range(1, 6):
            sheet.cell(row=start_row, column=column).font = Font(bold=True)

        current_row = start_row + 1
        for charge in booking.revenue_charges:
            sheet.append(
                [
                    "REVENUE",
                    charge.charge_category.value,
                    charge.description,
                    charge.container or "",
                    float(charge.amount.amount),
                ]
            )
            current_row += 1
        for charge in booking.cost_charges:
            sheet.append(
                [
                    "COST",
                    charge.charge_category.value,
                    charge.description,
                    charge.container or "",
                    float(charge.amount.amount),
                ]
            )
            current_row += 1

        for row in sheet.iter_rows(
            min_row=start_row + 1,
            max_row=max(current_row - 1, start_row + 1),
            min_col=5,
            max_col=5,
        ):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "#,##0.00"

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
