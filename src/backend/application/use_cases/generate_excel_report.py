"""Use case for exporting commission reports to Excel."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.application.dtos import CommissionReportRequest
from backend.application.use_cases.generate_commission_report import (
    GenerateCommissionReportUseCase,
)


class GenerateExcelReportUseCase:
    """Generate Excel bytes from commission report data."""

    def __init__(self, commission_report_use_case: GenerateCommissionReportUseCase) -> None:
        self.commission_report_use_case = commission_report_use_case

    def execute(self, request: CommissionReportRequest) -> bytes:
        """Return XLSX bytes for commission report data."""
        report = self.commission_report_use_case.execute(request)

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Commission Report"

        headers = [
            "Booking ID",
            "Client",
            "Created At",
            "Status",
            "Revenue",
            "Costs",
            "Margin",
            "Commission",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for item in report.items:
            sheet.append(
                [
                    item.booking_id,
                    item.client_name or "",
                    item.created_at.isoformat(),
                    item.status,
                    float(item.total_revenue),
                    float(item.total_costs),
                    float(item.margin),
                    float(item.commission),
                ]
            )

        sheet.append([])
        totals_row_index = sheet.max_row + 1
        sheet.append(
            [
                "TOTAL",
                "",
                "",
                "",
                float(report.totals.total_revenue),
                float(report.totals.total_costs),
                float(report.totals.total_margin),
                float(report.totals.total_commission),
            ]
        )
        for cell in sheet[totals_row_index]:
            cell.font = Font(bold=True)

        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=8):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "#,##0.00"

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
